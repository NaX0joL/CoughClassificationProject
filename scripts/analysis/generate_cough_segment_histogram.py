import ast
from math import ceil, isclose
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font



METADATA_PATH = Path("data/Elderly_Cough_Audio/metadata.xlsx")
OUTPUT_PATH = Path("outputs/elderly_cough_segment_duration_histograms.xlsx")
METADATA_WORKSHEET = "dynamo"
DETECTED_SEGMENTS_COLUMN = "DetectedCoughSegments"
DETECTED_SECONDS_COLUMN = "DetectedSeconds"
PATIENT_ID_COLUMN = "PatientID"
SEGMENT_FRAME_RATE_HZ = 16_000
RAW_AUDIO_SAMPLE_RATE_HZ = 48_000
HISTOGRAM_BIN_SECONDS = 0.5



def main() -> None:
    metadata_table = pd.read_excel(
        METADATA_PATH,
        sheet_name=METADATA_WORKSHEET,
        engine="openpyxl",
    )
    segment_table = _create_segment_table(metadata_table)
    histogram_table = _create_histogram_table(segment_table)
    _write_workbook(segment_table, histogram_table)
    return



def _create_segment_table(metadata_table:pd.DataFrame) -> pd.DataFrame:
    segment_rows = []

    for row_index, metadata_row in metadata_table.iterrows():
        frame_ranges = _parse_ranges(
            metadata_row[DETECTED_SEGMENTS_COLUMN],
            row_index,
            DETECTED_SEGMENTS_COLUMN,
        )
        second_ranges = _parse_ranges(
            metadata_row[DETECTED_SECONDS_COLUMN],
            row_index,
            DETECTED_SECONDS_COLUMN,
        )

        if len(frame_ranges) != len(second_ranges):
            raise ValueError(
                f"Segment and second range counts differ at metadata row {row_index}"
            )

        for segment_index, (frame_range, second_range) in enumerate(
            zip(frame_ranges, second_ranges),
            start=1,
        ):
            start_frame, end_frame = frame_range
            start_second, end_second = second_range
            _validate_frame_rate(
                start_frame,
                end_frame,
                start_second,
                end_second,
                row_index,
            )
            segment_rows.append(
                {
                    "patient_id": metadata_row[PATIENT_ID_COLUMN],
                    "segment_index": segment_index,
                    "start_frame": start_frame,
                    "end_frame": end_frame,
                    "duration_frames": end_frame - start_frame,
                    "start_seconds": start_second,
                    "end_seconds": end_second,
                    "duration_seconds": end_second - start_second,
                }
            )

    return pd.DataFrame(segment_rows)



def _parse_ranges(value:object, row_index:int, column_name:str) -> list[tuple[float, float]]:
    if pd.isna(value):
        return []

    try:
        ranges = ast.literal_eval(value)
    except (SyntaxError, ValueError) as error:
        raise ValueError(
            f"Invalid {column_name} value at metadata row {row_index}"
        ) from error

    return [(float(start), float(end)) for start, end in ranges]



def _validate_frame_rate(
    start_frame:float,
    end_frame:float,
    start_second:float,
    end_second:float,
    row_index:int,
) -> None:
    start_matches = isclose(
        start_frame / SEGMENT_FRAME_RATE_HZ,
        start_second,
        abs_tol=1e-9,
    )
    end_matches = isclose(
        end_frame / SEGMENT_FRAME_RATE_HZ,
        end_second,
        abs_tol=1e-9,
    )
    if not start_matches or not end_matches:
        raise ValueError(
            f"Segment frames do not match {SEGMENT_FRAME_RATE_HZ} Hz at metadata row {row_index}"
        )

    return



def _create_histogram_table(segment_table:pd.DataFrame) -> pd.DataFrame:
    maximum_duration = segment_table["duration_seconds"].max()
    bin_count = ceil((maximum_duration - HISTOGRAM_BIN_SECONDS) / HISTOGRAM_BIN_SECONDS)
    bin_edges = [
        HISTOGRAM_BIN_SECONDS + index * HISTOGRAM_BIN_SECONDS
        for index in range(bin_count + 1)
    ]
    histogram_rows = []

    for bin_index, lower_seconds in enumerate(bin_edges[:-1]):
        upper_seconds = bin_edges[bin_index + 1]
        is_final_bin = bin_index == len(bin_edges) - 2
        durations = segment_table["duration_seconds"]
        is_in_bin = durations.ge(lower_seconds) & durations.lt(upper_seconds)
        if is_final_bin:
            is_in_bin = durations.ge(lower_seconds) & durations.le(upper_seconds)

        histogram_rows.append(
            {
                "duration_bin_seconds": _format_bin(
                    lower_seconds,
                    upper_seconds,
                    is_final_bin,
                ),
                "start_frame": int(lower_seconds * SEGMENT_FRAME_RATE_HZ),
                "end_frame": int(upper_seconds * SEGMENT_FRAME_RATE_HZ),
                "segment_count": int(is_in_bin.sum()),
            }
        )

    return pd.DataFrame(histogram_rows)



def _format_bin(lower_seconds:float, upper_seconds:float, is_final_bin:bool) -> str:
    upper_operator = "≤" if is_final_bin else "<"
    return f"{lower_seconds:g}–{upper_operator}{upper_seconds:g} s"



def _write_workbook(
    segment_table:pd.DataFrame,
    histogram_table:pd.DataFrame,
) -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    duration_sheet = workbook.create_sheet("Segment durations")
    histogram_sheet = workbook.create_sheet("Histogram")

    _write_summary(summary_sheet, segment_table)
    _write_table(duration_sheet, segment_table)
    _write_table(histogram_sheet, histogram_table)
    _add_histogram_chart(histogram_sheet, len(histogram_table))

    workbook.save(OUTPUT_PATH)
    return



def _write_summary(summary_sheet, segment_table:pd.DataFrame) -> None:
    summary_rows = [
        ("Raw audio sample rate (Hz)", RAW_AUDIO_SAMPLE_RATE_HZ),
        ("Segment coordinate rate (frames/second)", SEGMENT_FRAME_RATE_HZ),
        ("Cough segments", len(segment_table)),
        ("Minimum duration (seconds)", segment_table["duration_seconds"].min()),
        ("Median duration (seconds)", segment_table["duration_seconds"].median()),
        ("Maximum duration (seconds)", segment_table["duration_seconds"].max()),
    ]
    summary_sheet.append(["Metric", "Value"])
    for summary_row in summary_rows:
        summary_sheet.append(summary_row)

    _style_header(summary_sheet)
    summary_sheet.column_dimensions["A"].width = 38
    summary_sheet.column_dimensions["B"].width = 18
    return



def _write_table(worksheet, table:pd.DataFrame) -> None:
    worksheet.append(table.columns.tolist())
    for row in table.itertuples(index=False, name=None):
        worksheet.append(row)

    _style_header(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        column_letter = column[0].column_letter
        worksheet.column_dimensions[column_letter].width = max(
            len(str(cell.value or "")) for cell in column
        ) + 2

    return



def _style_header(worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    return



def _add_histogram_chart(histogram_sheet, histogram_row_count:int) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Cough segment duration distribution"
    chart.y_axis.title = "Segments"
    chart.x_axis.title = "Duration (seconds)"
    chart.gapWidth = 0

    data = Reference(
        histogram_sheet,
        min_col=4,
        min_row=1,
        max_row=histogram_row_count + 1,
    )
    categories = Reference(
        histogram_sheet,
        min_col=1,
        min_row=2,
        max_row=histogram_row_count + 1,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    histogram_sheet.add_chart(chart, "F2")
    return



if __name__ == "__main__":
    main()
